import os
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook

from config import EXCEL_FILE


def salvar_excel(vagas):

    try:

        if os.path.exists(EXCEL_FILE):

            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        else:

            raise FileNotFoundError

    except (FileNotFoundError, BadZipFile):

        print("📄 Criando nova planilha Excel...")

        wb = Workbook()
        ws = wb.active

        ws.append([
            "Data",
            "Título",
            "Empresa",
            "Score",
            "Link"
        ])

    for vaga in vagas:

        ws.append([
            vaga["data"],
            vaga["titulo"],
            vaga["empresa"],
            vaga["score"],
            vaga["link"]
        ])

    wb.save(EXCEL_FILE)

    print("✅ Excel atualizado.")